import pandas as pd
import os
from openpyxl import load_workbook
import psutil

RAW_DATA_PATH_FILE = "data/raw_data.xlsx"
RESULT_PATH_FILE = "data/result.xlsx"
TEMPLATE_DATA_PATH_FILE = "data/template_data.txt"


def is_file_open(filepath):
    """ Kiểm tra xem file có đang mở hay không """
    try:
        # Kiểm tra file bằng cách thử mở và đóng nó
        with open(filepath, 'a'):
            return False
    except IOError:
        return True


def save_payloads_to_excel(payload_generator, filename=RAW_DATA_PATH_FILE, chunk_size=10000):
    # Kiểm tra nếu file đang mở
    if is_file_open(filename):
        print(f"File '{filename}' đang mở. Vui lòng đóng file và thử lại.")
        return

    # Mở một writer Excel để ghi từng payload vào
    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        chunk = []
        for i, payload in enumerate(payload_generator):
            chunk.append(payload)

            # Khi đủ chunk_size, ghi vào Excel
            if len(chunk) >= chunk_size:
                df = pd.DataFrame(chunk)
                df.to_excel(writer, sheet_name='Sheet1', startrow=i - len(chunk) + 1, header=(i == len(chunk) - 1),
                            index=False)
                chunk = []  # Reset chunk sau khi ghi

        # Ghi những phần còn lại nếu có (khi n không chia hết cho chunk_size)
        if chunk:
            df = pd.DataFrame(chunk)
            df.to_excel(writer, sheet_name='Sheet1', startrow=i - len(chunk) + 1, header=(i == len(chunk) - 1),
                        index=False)

    print(f"Dữ liệu đã được lưu vào tệp '{filename}'.")


def save_object_to_dict(obj):
    '''Chuyển đổi object thành dictionary'''
    try:
        # Sử dụng __dict__ để lấy thuộc tính của object dưới dạng dictionary
        return obj.__dict__
    except AttributeError:
        print("The object does not have a __dict__ attribute.")
        return {}


def save_response_to_excel(data, filename="response.xlsx"):
    '''Lưu dữ liệu vào file Excel sử dụng openpyxl với cột index'''
    try:
        # Nếu data là object, chuyển nó thành dictionary
        if hasattr(data, '__dict__'):
            data = save_object_to_dict(data)

        # Xử lý data nếu có giá trị là danh sách (list)
        for key, value in data.items():
            if isinstance(value, list):
                # Nếu giá trị là danh sách, chuyển thành chuỗi, các phần tử cách nhau bằng dấu phẩy
                data[key] = ", ".join(value)

        # Tạo DataFrame từ data đã xử lý
        df = pd.DataFrame([data])

        try:
            # Nếu file đã tồn tại, đọc nó
            workbook = load_workbook(filename)
            sheet = workbook.active

            # Lấy dòng cuối cùng có dữ liệu trong sheet hiện tại
            last_row = sheet.max_row

            # Append dữ liệu mới vào file Excel
            for idx, row in enumerate(df.values.tolist(), start=last_row + 1):
                # Thêm index vào đầu mỗi dòng dữ liệu
                sheet.append([idx] + row)  # thêm index vào đầu mỗi dòng

            workbook.save(filename)
            print(f"Data saved to {filename}")

        except FileNotFoundError:
            # Nếu file không tồn tại, tạo mới file Excel và thêm cột 'Index'
            df.insert(0, 'Index', range(1, len(df) + 1))  # Thêm cột Index vào DataFrame
            df.to_excel(filename, index=False)
            print(f"Data saved to {filename}")

    except Exception as e:
        print(f"Error saving data to Excel: {e}")


def open_excel_file(filename):
    '''Mở hoặc tạo mới file Excel'''
    try:
        # Kiểm tra xem file có tồn tại không, nếu không thì tạo mới
        try:
            # Mở workbook nếu tồn tại
            workbook = load_workbook(filename)
            print(f"Opened existing file: {filename}")
        except FileNotFoundError:
            # Tạo workbook mới nếu file không tồn tại
            workbook = pd.ExcelWriter(filename, engine='openpyxl')
            print(f"Created new file: {filename}")
    except Exception as e:
        print(f"Error opening file: {e}")
        raise e
    return workbook


def close_excel_file(workbook, filename):
    '''Đóng file Excel sau khi lưu xong'''
    try:
        workbook.save(filename)
        print(f"File {filename} saved and closed.")
    except Exception as e:
        print(f"Error closing file: {e}")


def is_file_open(filename):
    '''Kiểm tra xem file có đang mở không'''
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            # Kiểm tra nếu có một chương trình mở file này
            for item in proc.open_files():
                if item.path == filename:
                    return True
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass
    return False


def delete_file_if_open(filename):
    '''Nếu file đang mở, đóng nó và xóa'''
    if is_file_open(filename):
        print(f"File {filename} is open. Closing and deleting...")
        # Có thể thêm logic để đóng file nếu cần
        # Ví dụ: đóng ứng dụng hoặc làm việc với file đang mở thông qua giao diện người dùng hoặc hệ thống
        os.remove(filename)
    else:
        os.remove(filename)
        print(f"File {filename} deleted successfully.")
